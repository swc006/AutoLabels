import requests
import json
import docx2txt
from io import BytesIO
from dateutil.relativedelta import relativedelta
import pprint
import pandas
import pandas as pd
from zipfile import BadZipFile
from colorama import Fore
from docx import Document
import openpyxl 
import copy 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import datetime,timedelta
import pyautogui as pag
import time
import pyperclip
from PyPDF2 import PdfReader
import PySimpleGUI as sg
import sys
from PIL import Image
import os
import win32com.client

outlook = win32com.client.Dispatch('outlook.application')
mapi = outlook.GetNamespace("MAPI")
inbox = mapi.GetDefaultFolder(6)
messages = inbox.Items
received_dt = datetime.now() - timedelta(minutes = 6)
received_dt = received_dt.strftime('%m/%d/%Y %H:%M %p')
messages = messages.Restrict("[ReceivedTime] >= '" + received_dt + "'")
messages = messages.Restrict("@SQL=(urn:schemas:httpmail:subject LIKE '%Planner Extracted%')")

for account in mapi.Accounts: 
    print(account.DeliveryStore.DisplayName)

def openSAP(SAP_closed):
    """
    Open SAP if closed
    """
    
    xloc,yloc = pag.locateCenterOnScreen(SAP_closed, confidence=0.8)
    #print(xloc,yloc)
    pag.click(xloc,yloc)
    while not pag.locateCenterOnScreen(SAP_workspace, confidence=0.9):
        print('Waiting for workspace')
    x,y = pag.locateCenterOnScreen(SAP_workspace, confidence=0.9)
    pag.click(x,y)
    pag.hotkey('right'); pag.hotkey('down'); pag.hotkey('tab'); pag.hotkey('enter')
    while not pag.locateCenterOnScreen(SAP_Home, confidence=0.9):
        print('waiting for home')
        
def openSAP_wlogon(SAP_logonopen):
    """
    Open SAP and login
    """
    xloc,yloc = pag.locateCenterOnScreen(SAP_logonopen)
    pag.click(xloc,yloc)
    while not pag.locateCenterOnScreen(SAP_workspace, confidence=0.9):
        print('Waiting for workspace')
    x,y = pag.locateCenterOnScreen(SAP_workspace, confidence=0.9)
    pag.click(x,y); pag.hotkey('right'); pag.hotkey('down'); pag.hotkey('tab'); pag.hotkey('enter')
    while not pag.locateCenterOnScreen(SAP_Home, confidence=0.9):
        print('waiting for home')
    
def openSAP_home():
    """
    Go to SAP home
    """
    xloc,yloc = pag.locateCenterOnScreen(SAP_HomeLogon)
    pag.click(xloc,yloc)
    time.sleep(0.2)
    x,y = pag.locateCenterOnScreen(SAP_checkifhome)
    pag.click(SAP_checkifhome)
    while not pag.locateCenterOnScreen(SAP_Home, confidence=0.9):
        print('waiting for home')
    
def final_openSAP():
    """
    Finalized procedure for opening SAP
    """
    if pag.locateCenterOnScreen(SAP_closed, confidence=0.9):
        print('SAP closed')
        openSAP(SAP_closed)
    elif pag.locateCenterOnScreen(SAP_logonopen, confidence = 0.9):
        print('SAP Logon found')
        openSAP_wlogon(SAP_logonopen)
    elif pag.locateCenterOnScreen(SAP_HomeLogon, confidence=0.9):
        print('SAP fully open')
        openSAP_home()
        
def get_bearer():
    """
    Acquire bearer token from EQV veeva vault.
    """
    
    pag.click(x=60,y=30,button='right'); pag.hotkey('down'); pag.hotkey('enter'); time.sleep(0.1)
    pag.click(60,130); time.sleep(2)

    if pag.locateCenterOnScreen(veevauth, confidence=0.9):
        xloc,yloc = pag.locateCenterOnScreen(veevauth)
        pag.click(xloc,yloc)

    time.sleep(2)
    if pag.locateCenterOnScreen(cbutton, confidence=0.9):
        xloc,yloc = pag.locateCenterOnScreen(cbutton)
        pag.click(xloc,yloc)

    time.sleep(1)
    if pag.locateCenterOnScreen(cookie, confidence=0.9):
        xloc,yloc = pag.locateCenterOnScreen(cookie)
        pag.click(xloc,yloc)
    time.sleep(2)
    if pag.locateCenterOnScreen(eqv, confidence=0.9):
        xloc,yloc = pag.locateCenterOnScreen(eqv)
        pag.click(xloc,yloc)
        pag.click(xloc,yloc)

    pag.hotkey('down'); pag.hotkey('right')
    
    for i in range(5):
        pag.hotkey('down')

    if pag.locateCenterOnScreen(TK):
        xloc,yloc = pag.locateCenterOnScreen(TK)
        pag.click(xloc,yloc)

    else:
        pag.hotkey('down')
        if pag.locateCenterOnScreen(TK):
            xloc,yloc = pag.locateCenterOnScreen(TK)
            pag.click(xloc,yloc)

    pag.hotkey('tab'); pag.hotkey('tab')
    code = pag.hotkey('ctrl','c')
    bearer = pyperclip.paste()
    
    return bearer
              
def sendmail():
    """
    Send mail to start power automate flow.
    """
    mail = outlook.CreateItem(0)
    mail.To = account
    mail.Subject = 'Start Planner Extraction'
    mail.Send()

def sendandwait():
    """
    Send mail using sendmail() and wait for power automate to reply with planner data.
    """
    i = 0
    msg2 = []
    preplist = []
    sendmail()
    while i == 0:
        time.sleep(1)
        for msg in messages:
            if 'Planner Extracted' in str(msg):
                preppers = msg.body
                preplist = json.loads(preppers)
                i = 1
    new_preplist = [x.split('\n') for x in preplist]
    
    return new_preplist


def extractplanner(preplist,bearer):
    """
    Determine each preps PN, Volume, Storage, CIP, SIP, and Area.
    """
    
    cips = {'2003': {'1063':'20', '1060':'1', 'BRX-1063':'20', 'BRX-1060':'1'},
            '2008': {'1063':'22', '1060':'3', 'BRX-1063':'22', 'BRX-1060':'3'},
            '2072': {'1520':'132', '1128':'132', '1132':'132', '1186':'132', '2150':'132',
                     'smform':'133'},
            '2075': {'1128':'130', '1132':'130', '1186':'130', '2150':'130', '1520':'130',
                     'smform':'131',
                     '1151':'152', '1154':'152', '1513':'152', 
                     '1151_2':'156', '1154_2':'156',
                     'lgform':'168', 
                     'TP1126':'172', '1128':'172', '1132':'172'},
            '2078': {'1154': '161', '1151':'161', '1513':'161',
                     'lgform':'169',
                     'TP1238':'177', '1520':'177', '1128':'177', '1132':'177', '2150':'177'},
            '2081': {'1163':'138', '1166':'138', '1169':'138', '1172':'138',
                     '1157':'147', '1160':'147',
                     '1151':'167', '1154':'167', '1513':'167',
                     '1466':'385',
                     '4018':'640', '4019':'640', '4020':'640'},
            '2084': {'1163':'142', '1166':'142', '1169':'142', '1172':'142', 
                     '1157':'150', '1160':'150', 
                     '1466':'384', 
                     '4018':'641', '4019':'641', '4020':'641'}}

    cips_hold = {'1151': {'lgform':'93', 'lgpure':'96', 'smpure':'102'},
                 '1154': {'lgpure':'98', 'smpure':'104'},
                 '1157': {'1157':'83', 'lgpure':'85', 'lgform':'86', 'smpure':'87', 'smform':'88'},
                 '1160': {'1160':'84', 'lgpure':'89', 'smpure':'90'},
                 '1163': {'lgpure':'72', 'lgform':'80'},
                 '1166': {'lgpure':'70', 'lgform':'74'},
                 '1169': {'lgpure':'71'},
                 '1172': {'lgpure':'73'},
                 '1513': {'smpure':'106', 'smform':'109'}}
                 
    # SIPS
    sips = {'1163': {'lgpure':'245', 'lgform':'316'},
            '1166': {'lgpure':'317', 'lgform':'246'},
            '1169': {'lgpure':'247'},
            '1172': {'lgpure':'248'},
            '1151': {'lgpure':'307', 'lgform':'308', 'smpure':'309'},
            '1154': {'lgpure':'305', 'smpure':'305'},
            '1513': {'smform':'303', 'smform':'304'}}
            
    LForm_SIP = {'TK1163':'245', 'TK1166':'246', 'TK1151':'308'}
    LPure_SIP = {'TK1163':'316', 'TK1166':'317', 'TK1169':'247', 'TK1172':'248', 'TK1151':'307', 'TK1154':'305'}

    SForm_SIP = {'TK1513':'303'}
    SPure_SIP = {'TK1513':'304', 'TK1154':'306', 'TK1151':'309'}
    
    endpoint = "hidden'
    sprsearch = "hidden"
    heada2 = {"Authorization":"Bearer "+bearer, "Accept":"application/json"}
    
    bag_pns = ['4102464','4000076','bags','Bags']
    preptank_ids = ['MXT','2081','2084','2075','2072','2078','2008','2003','Bottle','mxt','bottle']
    holdtank_ids = ['1157','1160','1151','1154','1513','1172','1163','1169','1063']
    porttank_ids = ['1128','1132','1186','2150','1520']
    
    spr_codes = ['A0','A1','A2','A3','A4','A5','A6','A7','A8','A9','80','MABR','AZDoc','A005148','A0014D','A0191-D']
    not_spr_codes = ['PS80','800L']
    
    preps_list = []
    to_fixes = {'1x 1':'1x1', ' x ':'x', 'x ': 'x', ' x':'x', 'x ':'x',
                ',1':', 1', '/ ': '/', ')M':') M', '(':' (', 'ml':'mL',
                'L1':'', 'e2 L':'e 2L','0in':'0 in', ')T':') T', '0(':'0 (',
                '1(':'1 (', '2(':'2 (', '3(':'3 (', '4(':'4 (', '5(':'5 (',
                '6(':'6 (','7(':'7 (', '8(':'8 (', '9(':'9 (', 'L(S':'L (S',
                'mL(B': 'mL (B', 'to1':'to 1', 'to2':'to 2', 'to3':'to 3',
                'to4':'to 4', 'to5':'to 5', 'to6':'to 6', 'to7':'to 7',
                'to8':'to 8','to9':'to 9', '1in':'1 in', '2in':'2 in',
                '3in':'3 in', '4in':'4 in', '5in':'5 in', '6in':'6 in',
                '7in':'7 in', '8in':'8 in', '9in':'9 in', '1to':'1 to',
                '2to':'2 to', '3to':'3 to', '4to':'4 to', '5to':'5 to',
                '6to':'6 to', '7to':'7 to', '8to':'8 to', '9to':'9 to',
                '0 L':'0L', '1 L':'1L', '2 L':'2L', '3 L':'3L', '4 L':'4L',
                '5 L':'5L', '6 L':'6L', '7 L':'7L', '8 L':'8L', '9 L':'9L',
                '0 mL':'0mL', '1 mL':'1mL', '2 mL':'2mL', '3 mL':'3mL',
                '4 mL':'4mL', '5 mL':'5mL', '6 mL':'6mL', '7 mL':'7mL',
                '8 mL':'8mL', '9 mL':'9mL', 'Lb':'L b'}

    total_preplist = []
    total_prepdict = {}
    total_prepdict_final = {}
    total_prepdict_final2 = {}
    count = -1
    
    for day in preplist:
        preplist_enh = []
        for string in day:
            prepstring = []
            
            for key,val in to_fixes.items():
                if key in string:
                    string = string.replace(key,val)
                    prepstring.append(string)
            else:
                preplist_enh.append(string)
                continue
        total_preplist.append(preplist_enh)
    
    pn_prep_dict = {}
    for day in total_preplist: # Preplist is the formatted information from the current planner day for each prep
        pn_dict = {}
        index_counter = -1
        preps_for_day = []
        preps_for_day_total = []
        indices = []
        iteration = 0
        newlist = []
        indices_new = []
        day_string = " ".join(day)
        day_split = day_string.split(' ')
        prepfinal = []
        
        for index,elem in enumerate(day_split): # Get index of preps
            for code in spr_codes:
                if code in elem and 'PS' not in elem and '800L' not in elem and '80L' not in elem:
                    indices.append(index)
                    break
                
        if len(indices) == 1:
            prepfinal.append(day_string)
        elif len(indices) == 2:
            prepfinal.append(" ".join(day_split[0:indices[1]]))
            prepfinal.append(" ".join(day_split[indices[1]:]))
        elif len(indices) == 3:
            prepfinal.append(" ".join(day_split[0:indices[1]]))
            prepfinal.append(" ".join(day_split[indices[1]:indices[2]]))
            prepfinal.append(" ".join(day_split[indices[2]:]))

        for item in prepfinal:
            count += 1
            pn = ''
            prepvessel = ''
            volume = ''
            area = ''
            pn_pn = ''
            prep_PN = ''
            storagelist = []
            storagelist_final = []
            sublot_check = 0
            donot_count = 0
            
            # AREA #
            if 'Sm PA' in item or 'sm PA' in item or 'SM PA' in item or 'Sm Pa' in item or 'smPa' in item:
                area = 'smpure'
            elif 'Lg PA' in item or 'lg PA' in item or 'LG PA' in item or 'Lg Pa' in item or 'lgPa' in item:
                area = 'lgpure'
            elif 'Lg FM' in item or 'lg FM' in item or 'LG FM' in item or 'Lg Fm' in item or 'lgfm' in item:
                area = 'lgform'
            elif 'Sm FM' in item or 'sm FM' in item or 'SM FM' in item or 'Sm Fm' in item or 'smfm' in item:
                area ='smform'
            if 'bag of' in item or 'bags of' in item or 'bags 0f' in item or 'bags Of' in item or 'Bags Of' in item:
                donot_count = 1
                
            # PN #
            prep_split = item.split(' ')
            for elem in item.split(' '):
                if 'sublot' in item or 'Sublot' in item:
                    if elem.startswith('(') and elem.endswith('L'):
                        volume = elem[1:]
                for sprcode in spr_codes:
                    if sprcode in elem and '80L' not in elem and '800L' not in elem:
                        if '/' in elem:
                            split0 = elem.split('/')[0]
                            pn_pn = split0
                            split1 = elem.split('/')[1]
                            if 'MABR' in split0:
                                pn = split0
                                break
                            elif 'MABR' in split1:
                                pn = split1
                                break
                            if '80' in split0 and 'MABR' not in split0 and 'A' not in split0 and 'PS' not in split0:
                                pn = split0
                                break
                            elif '80' in split1 and 'MABR' not in split1 and 'A' not in split1 and 'PS' not in split1:
                                pn = split1 
                                break
                            if 'A' in split0 and 'MABR' not in split0:
                                pn = split0
                                break
                            elif 'A' in split1 and 'MABR' not in split1:
                                pn = split1
                                break
                        else:
                            if 'PS' not in elem:
                                pn = elem
                                pn_pn = elem

            if len(pn) == 6 and pn.endswith('D'):
                pn_pn = pn_pn[0:5] +'-'+pn_pn[5]            
            elif len(pn) == 7 and pn.startswith('A'):
                pn_pn = pn_pn[0]+'-'+pn_pn[1:]
                
            if len(pn) == 6 and pn.endswith('D'):
                pn = pn[0:5] +'-'+pn[5]
            elif len(pn) == 7 and pn.startswith('A'):
                pn = pn[0]+'-'+pn[1:]
                
            if len(pn_pn) == 7 and pn_pn.startswith('A'):
                pn_pn = pn_pn[0]+'-'+pn_pn[1:]
                
            if pn == 'A0643':
                pn = 'A-0643'
            if pn == 'A0191D':
                pn = 'A0191-D'
            if pn == 'A0101D':
                pn = 'A0101-D'
            if pn == 'A0014D':
                pn = 'A0014-D'
            if pn_pn == 'A0014D':
                pn_pn = 'A0014-D'
            if pn_pn == 'A0101D':
                pn_pn = 'A0101-D'
            if pn_pn == 'A-0013-D':
                pn_pn = 'A0013-D'
            if pn_pn == 'A0191D':
                pn_pn = 'A0191-D'
            if pn == 'A0647':
                pn = 'A-0647'

            # PREP VESSEL #
            for elem in item.split(' '):
                for prepv in preptank_ids: # TANK
                    if prepv in elem and 'bottles' not in elem:
                        prepvessel = elem
            
            # STORAGE VESSELS #
            for elem in item.split(' '): 
                if ',' in elem:
                    elem = elem.replace(',','')
                if 'x' in elem and 'c' not in elem and 'fle' not in elem:
                    storagelist.append([elem.split('x')[0],elem.split('x')[1]])
                
                for holdtank in holdtank_ids:
                    if holdtank in elem:
                        if ')' in elem:
                            elem = elem.replace(')','')
                        storagelist.append(elem)
                
                for porttank in porttank_ids:
                    if porttank in elem:
                        if ')' in elem:
                            elem = elem.replace(')','')
                        storagelist.append(elem)
            if 'in 4000076' in item:
                storagelist.append(['1','2L'])
            if 'to 50L' in item:
                storagelist.append(['1','50L'])
            if 'to 100L' in item:
                storagelist.append(['1','100L'])
            if 'to Bag' in item or 'to bag' in item:
                if ',' in item:
                    item = item.replace(',','')
                storagelist.append(['1','Bag'])
            
            for word in storagelist:
                if 'A0' in word or 'MABR' in word or 'AZ' in word:
                    pass
                else:
                    storagelist_final.append(word)
            
            if 'sublot' in item or 'Sublot' in item:
                sublot_check = 1
            if 'hyclone' in item or 'hyclones' in item:
                hydroxide_stor = 'hyclone'
            elif 'stedim' in item:
                hydroxide_stor = 'stedim'
            else:
                hydroxide_stor = 'tank'
            
            for v in range(len(prep_split)):
                if '(' in prep_split[v]:
                    try:
                        if 'in' in prep_split[v+1]:
                            volume = prep_split[v].split('(')[1]
                    except IndexError:
                        break
                        
            for v in range(len(prep_split)):
                if prep_split[v].startswith('(') and prep_split[v].endswith(')'):
                    volume = prep_split[v].split('(')[1].split(')')[0]
                    
            pn_prep_dict = {'PN':pn, 'Area':area, 'Prep Vessel':prepvessel,
                            'Storage':storagelist_final, 'NaOH Store':hydroxide_stor,
                            'SAP PN':pn_pn, 'Volume':volume, 'Sublot?':sublot_check}
            
            total_prepdict[count] = pn_prep_dict
    
    for key,val in total_prepdict.items():
        if val['Prep Vessel'] != '':
            total_prepdict_final[key] = val
        elif val['Prep Vessel'] == '' and val['Sublot?'] == 1:
            total_prepdict_final[key] = val
            
    for key,val in total_prepdict_final.items():
        cip_touse = []
        siptouse = []
        use_portables = False
        use_holds = False
        use_bags = False
        cip_touse = []
        
        # No CIP needed
        if val['Prep Vessel'] == 'Bottle' or val['Prep Vessel'] == 'bottle' or val['Prep Vessel'] == 'MXT' or val['Prep Vessel'] == 'mxt':
            cip = ''
        else: 
            for item in val['Storage']:
                if type(item) == str and item in porttank_ids:
                    pass
                elif type(item) == str and item in holdtank_ids:
                    try:
                        cip = cips[val['Prep Vessel']][item]
                        if cip not in cip_touse:
                            cip_touse.append(cip)
                    except KeyError:
                        cip = ''
                        print('CIP from ',val['Prep Vessel'],' to Hold Tank/TP not found')
                else:
                    cip_touse.append('any')
                if len(val['Area']) >= 1:
                    if item in holdtank_ids:
                        try:
                            cip_hold = cips_hold[item][val['Area']]
                            cip_touse.append(cip_hold)
                        except KeyError:
                            print('CIP not found')

                        try: 
                            sip_hold = sips[item][val['Area']]
                            siptouse.append(sip_hold)
                        except KeyError:
                            print('SIP not found')

        if 'any' in cip_touse and len(cip_touse) > 1:
            cip_touse.remove('any')

        total_prepdict_final2[key] = {'PN':val['PN'], 'Area':val['Area'], 'Prep Vessel':val['Prep Vessel'], 'Storage':val['Storage'], 'CIP':cip_touse, 'SIP':siptouse, 'NaOH Store':val['NaOH Store'], 'SAP PN':val['SAP PN'], 'Volume':val['Volume'], 'Sublot?':val['Sublot?']}
    for key,val in total_prepdict_final2.items():
        if val['Sublot?'] == 1:
            total_prepdict_final2[key]['CIP'] = ''
        if val['Storage'] == []:
            total_prepdict_final2[key]['Storage'] = ['1','200L']
        if val['Prep Vessel'] == 'MXT' or val['Prep Vessel'] == 'mxt' or val['Prep Vessel'] == 'Bottle' or val['Prep Vessel'] == 'bottle':
            total_prepdict_final2[key]['CIP'] = ''
    
    return total_prepdict_final2

def goto_COOISPI(): 
    """
    Navigate to COOISPI from Main screen.
    """
    if pag.locateCenterOnScreen(SAP_COOISPI, confidence=0.9):
        x,y = pag.locateCenterOnScreen(SAP_COOISPI)
        pag.click(x,y)
        pag.click(x,y)
    time.sleep(4)       
    pag.typewrite('o'); time.sleep(0.5)
    pag.hotkey('enter')

    time.sleep(1)
    for i in range(7):
        pag.hotkey('tab')
    time.sleep(1) # Stop at Material

def COOISPI_2(pns_to_lookup_prep): 
    """
    Get our POs and BNs.
    """
    time.sleep(3)
    count = 0

    for key,val in pns_to_lookup_prep.items(): # Start at material
        SAP_PO_list = []
        spec_po = []
        po_l = []
        PO = 'error'
        BN = 'error'

        if count >= 1:
            for i in range(7):
                pag.hotkey('tab')

        if 'MABR' in val['PN']:
            pag.typewrite(val['SAP PN'])
        else:
            pag.typewrite(val['PN'])

        if count == 0:
            for i in range(73):
                pag.hotkey('tab')

            pag.typewrite('5') # Get most recent POs
        pag.hotkey('F8')
        time.sleep(4)
        pag.click(596, 235)
        time.sleep(1.7)
        pag.hotkey('ctrl','c')
        list_of_pos = pyperclip.paste()
        pos_split = list_of_pos.split('\t')

        for word in pos_split: # Split this up
            if '\n' in word:
                word = word.split('\n')
                SAP_PO_list.append(word[0])
                SAP_PO_list.append(word[1])
            else:
                SAP_PO_list.append(word)
        SAP_PO_list.pop(-1)

        composite_list = [SAP_PO_list[x:x+17] for x in range(0, len(SAP_PO_list),17)] # Convert to list of each PO in SAP

        for prep_po in composite_list:
            if 'CNF' not in prep_po[6]: # Get the non completed PO
                if prep_po[8] == '0' and prep_po[9] == '0':
                    po_l.append(prep_po)
                    spec_po = prep_po

        if len(po_l) == 1:
            spec_po = po_l[0]
        else:
            for item in po_l:
                if 'TECO' not in item[6]:
                    try:
                        if len(val['Volume']) >= 1:
                            if val['Volume'].endswith('L'):
                                int_vol = val['Volume'].split('L')[0]
                                int_sap_vol = int(spec_po[7].replace(',',''))
                                if int_vol/int_sap_vol in range(0.75,1.25):
                                    
                                    spec_po = item
                    except Exception:          
                        spec_po = item
                    break
        PO = spec_po[3]
        BN = spec_po[5]
        SAP_PN = spec_po[1]
        
        val['SAP PN'] = SAP_PN
        val['PO'] = PO
        val['BN'] = BN
        
        pag.hotkey('F3')
        time.sleep(1)
        count += 1

    return pns_to_lookup_prep

def getsprid(pns_to_lookup,bearer): # Get the document id given the spr PN or doc code
    
    endpoint = 'hidden'
    sprsearch = 'hidden'
    heada2 = {"Authorization":"Bearer "+bearer,
         "Accept":"application/json"}
    sprid = {}
    sprname = {}
    sprid2 = {}
    doc_name = ""
 
    for key,val in pns_to_lookup.items():
        sprcontents = requests.get(endpoint + sprsearch + '"'+val['PN']+'" Master Batch Record',headers=heada2)
        sprcontents_dict = json.loads(sprcontents.text)

        try:
            previous_doc_no = sprcontents_dict['documents'][0]['document']['previous_document_number__c']
        except Exception:
            previous_doc_no = 'None'
        if len(sprcontents_dict) > 1:
            try:
                doc_id = sprcontents_dict['documents'][0]['document']['id'] 
            except IndexError:
                print(sprcontents_dict)
            doc_name = sprcontents_dict['documents'][0]['document']['name__v']
        else:
            doc_id = sprcontents_dict['documents']['document']['id'] 
            doc_name = sprcontents_dict['documents']['document']['name__v']
        print(Fore.BLUE+"",doc_name)
        
        if '(GPFN) / ' in doc_name:
            doc_name = doc_name.split('(GPFN) / ')[1]
        elif '(GPF-N): ' in doc_name:
            doc_name = doc_name.split('(GPF-N): ')[1]
        elif '(GPFN):' in doc_name:
            doc_name = doc_name.split('(GPFN): ')[1]
        elif '(GPF-N) ' in doc_name:
            doc_name = doc_name.split('(GPF-N) ')[1]
        elif '(GPFN) ' in doc_name:
            doc_name = doc_name.split('(GPFN) ')[1]
        
        elif 'GPF-N PN: ' in doc_name:
            doc_name = doc_name.split('GPF-N PN: ')[1]
            
        elif '(GPF) PN: ' in doc_name:
            doc_name = doc_name.split('(GPF) PN: ')[1]
        elif 'GPF PN: ' in doc_name:
            doc_name = doc_name.split('GPF PN: ')[1]
        elif '(GPF) ' in doc_name:
            doc_name = doc_name.split('(GPF) ')[1]
        elif ' / ' in doc_name:
            doc_name = doc_name.split(' / ')[1]
        
        if 'GPF PN: ' in doc_name:
            doc_name = doc_name.replace('GPF PN: ',' ')
        if ' / ' in doc_name:
            doc_name = doc_name.split(' / ')[1]
        if 'Benzyl A' in doc_name and 'Benzyl Alcohol' not in doc_name:
            doc_name = doc_name.replace('Benzyl A','Benzyl Alcohol')
        doc_split = doc_name.split(' ')
        if 'PN' in doc_split[0] and doc_split[1].startswith('80'):
            doc_split.pop(0)
            doc_split.pop(0)
            
        if '-' in doc_split[0] or doc_split[0] == '/':
            doc_split.pop(0)
            
        if doc_split[0].startswith('80') and doc_split[0].endswith('/'):
            doc_split.pop(0)

        sprid[val['PN']] = " ".join(doc_split)+'@'+str(previous_doc_no)+'@'+str(doc_id)+r'/file'
    print(sprid)
    return sprid

def decodespr(sprveevaid,bearer,pns_to_lookup_prep): # Get the contents of the SPR given the document id from getsprid()
    endpoint = 'hidden'
    doccheck = 'hidden'
    heada2 = {"Authorization":"Bearer "+bearer,
         "Accept":"application/json"}
    pn_sprcontents = {}
    sprveevaid_new = {}
    
    hazard_info = pandas.read_excel(r'hidden') # Excel sheet location
    df = hazard_info.to_dict(orient='dict') # Convert to DataFrame
    spr_folder = 'hidden'
    filelist = [f for f in os.listdir(spr_folder)]
   
    hazard_dict = {}
    
    for keya,vala in df['Component'].items():
        typeof = ''
        for keyb,valb in df['Hazards'].items():
            if keya==keyb:
                hazard_dict[vala]=valb
    
    count = -1
    for key,val in sprveevaid.items():
        count += 1
        print(key,val)
        haz_list = ''
        labelvalues = {'PN':'',"pH":0,"osmo":0,"cond":0,"turb":0,"dark":0,"temp":"","expiration":[],'hazards':''}
        
        if val.split('@')[2] == '1119985/file':
            sprcontents = requests.get(endpoint + doccheck + '1143755/file',headers=heada2)
        elif val.split('@')[2] == '1119982/file':
            sprcontents = requests.get(endpoint + doccheck + '1119983/file',headers=heada2)
        elif val.split('@')[2] == '658663/file':
            sprcontents = requests.get(endpoint + doccheck + '658663/file',headers=heada2)
        else:
            sprcontents = requests.get(endpoint + doccheck + val.split('@')[2],headers=heada2)
            
        
        temp_holder = ''
        hazards = ''
        
        docx = BytesIO(sprcontents.content)
        try:
            docx_total = docx2txt.process(docx)
            typeof = 'docx'
        except Exception:
            docx_total = str(sprcontents.content)
            typeof = 'doc?'
        
        if r':\t' in docx_total:
            docx_total = docx_total.replace(r':\t',' ')
        if r'\x07' in docx_total:
            docx_total = docx_total.replace(r'\x07',' ')
        if r'\r\r\r\r\r\r\r\r\r\r' in docx_total:
            docx_total = docx_total.replace(r'\r\r\r\r\r\r\r\r\r\r',' ')
        if r'\r' in docx_total:
            docx_total = docx_total.replace(r'\r',' ')
        if r'\x01\x15' in docx_total:
            docx_total = docx_total.replace(r'\x01\x15',' ')
        if r'\x13' in docx_total:
            docx_total = docx_total.replace(r'\x13',' ')
        if r'\xb0c' in docx_total:
            docx_total = docx_total.replace(r'\xb0c',' ')
        if r'\x96' in docx_total:
            docx_total = docx_total.replace(r'\x96','-')
        if r'\xb0' in docx_total:
            docx_total = docx_total.replace(r'\xb0','°')
            
        if '/' in val.split('@')[0]:
            valsplit1 = val.split('@')[0].replace('/','')
        else:
            valsplit1 = val.split('@')[0]

        for files in filelist:
            if valsplit1 in files:
                print('SPR already saved')
                break
        else:
            with open(spr_folder + '\\'+valsplit1.replace(':','') +'.txt','w+',encoding='UTF-8') as f:
                try:
                    f.write(docx_total)
                except UnicodeEncodeError:
                    print('Unicode Error?')
            
        # Get EXP
        if r'\x00Description' in docx_total:
            desc = docx_total.split(r'\x00Description')[1].split('Solution Transfer')[0]
            typeof += r'\xoo'
        else:
            desc = docx_total.lower().split(r'description')[1].split('solution transfer')[0]
            typeof += 'norm'
        
        
        if '15 - 25' in docx_total or '15-25' in docx_total or '15- 25' in docx_total or '15 -25' in docx_total or '15 – 25' in docx_total:
            labelvalues['temp'] = '15 - 25'
        elif '2 - 8' in docx_total or '2-8' in docx_total or '2- 8' in docx_total or '2 -8' in docx_total or '2 – 8' in docx_total:
            labelvalues['temp'] = '2 - 8'
        if 'final ph' in docx_total.lower():
            labelvalues['pH'] = 1
        if 'final conductivity' in docx_total.lower():
            labelvalues['cond'] = 1
        if 'final turb' in docx_total.lower():
            labelvalues['turb'] = 1
        if 'final osm' in docx_total.lower():
            labelvalues['osmo'] = 1
        if 'dark' in docx_total.lower():
            labelvalues['dark'] = 1
            
        try:
            for_exp = docx_total.split('load cells')[0].split('up to')[1]
            
        except IndexError:
            for_exp = docx_total.lower().split('sublot solution')[0].split('special instructions')[1]
        
        
        if '90 days in a hyclone bag' in for_exp.lower(): # If hydroxide prep choose which expiration to use
            for key3,val3 in pns_to_lookup_prep.items():
                if val3['PN'] == key:
                    if val3['NaOH Store'] == 'hyclone':
                        labelvalues['expiration'] = 90
                        break
                    elif val3['NaOH Store'] == 'stedim':
                        labelvalues['expiration'] = 7
                        break
                    else:
                        labelvalues['expiration'] = 180
                        break
        else: # If not hydroxide prep
            if 'days' in for_exp:
                labelvalues['expiration'] = int(for_exp.split('days')[0])
            elif 'months' in for_exp:
                labelvalues['expiration'] = str(int(for_exp.split('months')[0])) + 'months'
            elif 'hours' in for_exp:
                print(for_exp)
                labelvalues['expiration'] = str(int(for_exp.split('hours')[0])) + 'hours'
        for key2,val2 in hazard_dict.items():
            if key2.lower() in docx_total.lower():
                if val2 not in haz_list: # No duplicate hazards
                    haz_list += val2

        if 'May' in haz_list or 'Causes' in haz_list or 'eye' in haz_list or 'resp' in haz_list or 'Eye' in haz_list or 'skin' in haz_list or 'Skin' in haz_list:
            if 'NO HAZARDS/PELIGROS.' in haz_list: # If other components have hazards and one does not, remove the non hazard
                haz_list = haz_list.replace('NO HAZARDS/PELIGROS.','')
        
        haz_list = haz_list.replace('.','. ')
        labelvalues['hazards'] = haz_list[:-1]
        
        labelvalues['PN'] = key
        pn_sprcontents[count] = labelvalues
    
    return pn_sprcontents

def createlabelstart(pns_to_lookup_prep):    
    """
    Create labels for each prep if they require them.
    """
    
    workbook = load_workbook(filename=r"hidden") 
    sheet = workbook.active 
    totalnumofvessels = 0
    totaltypesofvessels = 0
    labelcount = -1
    thin_border = Border(bottom=Side(style='thin'))
    
    count = 0
    
    for key,val in pns_to_lookup_prep.items():
        uses_hours = 0
        count_to_fpage = 5
        count += 1
        haz_img_list = []
        dont_inc_vnum = 0
        haz_img_list.append(filler_img) # Create a temporary hazard image to put into the sheet
        if 'corrosive' in val['hazards'] or 'Corrosive' in val['hazards']:
            haz_img_list.append(corrosive_img)
        if 'irritation' in val['hazards'] or 'irritant' in val['hazards'] or 'Irritation' in val['hazards']:
            haz_img_list.append(harmful_img)
        if 'flammable' in val['hazards'] or 'Flammable' in val['hazards']:
            haz_img_list.append(flammable_img)
        if 'toxic' in val['hazards'] or 'Toxic' in val['hazards']:
            haz_img_list.append(toxic_img)
        if 'aquatic' in val['hazards'] or 'Aquatic' in val['hazards']:
            haz_img_list.append(environmental_img)
            
        images = [Image.open(x) for x in haz_img_list]
        widths, heights = zip(*(i.size for i in images))

        total_width = sum(widths)
        max_height = max(heights)

        new_im = Image.new('RGB', (total_width, max_height))
        x_offset = 0
        for im in images:
            new_im.paste(im, (x_offset,0))
            x_offset += im.size[0]
        new_im.save(r'hidden'+str(count)+'.png')
        
        
        for item in val['Storage']:
            if type(item) == str:
                break
            if len(item) == 1:
                break
            else:
                totaltypesofvessels += 1
                
                for i in range(int(item[0])+1):
                    count_to_fpage -= 1
                    totalnumofvessels += 1
                    labelcount += 1
                    thin = Side(border_style="thin", color="000000")
                    sheet.merge_cells('A'+str(1+(8*(labelcount)))+':'+'B'+str(1+(8*(labelcount))))
                    sheet.merge_cells('A'+str(2+(8*(labelcount)))+':'+'B'+str(2+(8*(labelcount))))
                    sheet.merge_cells('A'+str(3+(8*(labelcount)))+':'+'B'+str(3+(8*(labelcount))))
                    sheet.merge_cells('A'+str(4+(8*(labelcount)))+':'+'B'+str(4+(8*(labelcount))))
                    sheet.merge_cells('A'+str(5+(8*(labelcount)))+':'+'B'+str(5+(8*(labelcount))))
                    sheet.merge_cells('A'+str(6+(8*(labelcount)))+':'+'B'+str(6+(8*(labelcount))))
                    
                    
                    sheet["B"+str(2+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["B"+str(3+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["B"+str(4+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["B"+str(5+(8*(labelcount)))].border = Border(bottom=thin)
                    
                    sheet["C"+str(2+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["D"+str(2+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["E"+str(2+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["F"+str(2+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["G"+str(2+(8*(labelcount)))].border = Border(bottom=thin)
                    
                    sheet["C"+str(3+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["D"+str(3+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["E"+str(3+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["F"+str(3+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["G"+str(3+(8*(labelcount)))].border = Border(bottom=thin)
                    
                    sheet.merge_cells('C'+str(1+(8*(labelcount)))+':'+'G'+str(1+(8*(labelcount))))
                    sheet.merge_cells('C'+str(2+(8*(labelcount)))+':'+'G'+str(2+(8*(labelcount))))
                    sheet.merge_cells('C'+str(4+(8*(labelcount)))+':'+'G'+str(6+(8*(labelcount))))

                    sheet.merge_cells('C'+str(4+(8*(labelcount)))+':'+'C'+str(6+(8*(labelcount))))
                    sheet["C"+str(4+(8*(labelcount)))].alignment = Alignment(wrap_text = True,indent = 2.0)
                    rd = sheet.row_dimensions[8+(8*(labelcount))]
                    rd.height = 43.0
                    
                    ## TITLE ##
                    sheet["A"+str(1+(8*(labelcount)))].fill = PatternFill("solid", fgColor="0000FF")
                    sheet["A"+str(1+(8*(labelcount)))] = 'In-Process/Equipment Status'
                    sheet["A"+str(1+(8*(labelcount)))].font  = Font(b=True, color="FFFFFF",size=14)
                    sheet["A"+str(1+(8*(labelcount)))].alignment = Alignment(horizontal = 'center')

                    sheet["C"+str(1+(8*(labelcount)))].fill = PatternFill("solid", fgColor="FF0000")
                    sheet["C"+str(1+(8*(labelcount)))] = 'HAZARD INFORMATION'
                    sheet["C"+str(1+(8*(labelcount)))].font  = Font(b=True, color="000000",size=14)
                    sheet["C"+str(1+(8*(labelcount)))].alignment = Alignment(horizontal = 'center')

                    ## DESCRIPTION ##
                    sheet["A"+str(2+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["C"+str(2+(8*(labelcount)))].border = Border(bottom=thin)

                    if len(val['description']) < 52: 
                        sheet["A"+str(2+(8*(labelcount)))].font  = Font(b=True, color="000000",size=9)
                        sheet["C"+str(2+(8*(labelcount)))].font  = Font(b=True, color="000000",size=9)
                    elif len(val['description']) >= 70: 
                        sheet["A"+str(2+(8*(labelcount)))].font  = Font(b=True, color="000000",size=7)
                        sheet["C"+str(2+(8*(labelcount)))].font  = Font(b=True, color="000000",size=7)
                    elif len(val['description']) >= 52: 
                        sheet["A"+str(2+(8*(labelcount)))].font  = Font(b=True, color="000000",size=8)
                        sheet["C"+str(2+(8*(labelcount)))].font  = Font(b=True, color="000000",size=8)

                    sheet["A"+str(2+(8*(labelcount)))] = 'Description: ' + val['description']
                    sheet["C"+str(2+(8*(labelcount)))].alignment = Alignment(indent = 2.0)
                    sheet["C"+str(2+(8*(labelcount)))] = 'Description: ' + val['description']

                    ## PN BN ##
                    if val['PN'] != val['SAP PN']:
                        sheet["A"+str(3+(8*(labelcount)))] = 'PN: ' + val['SAP PN']+'/'+val['PN'] + '                               BN: ' + val['BN']
                    else:
                        sheet["A"+str(3+(8*(labelcount)))] = 'PN: ' + val['PN'] + '                               BN: ' + val['BN']
                    sheet["A"+str(3+(8*(labelcount)))].font = Font(bold=True,size=10)
                    sheet["A"+str(3+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["C"+str(3+(8*(labelcount)))].alignment = Alignment(indent = 2.0)
                    if val['PN'] != val['SAP PN']:
                        sheet["C"+str(3+(8*(labelcount)))] = 'PN: ' + val['SAP PN']+'/'+val['PN'] + '                               BN: ' + val['BN']
                    else:
                        sheet["C"+str(3+(8*(labelcount)))] = 'PN: ' + val['PN'] + '                               BN: ' + val['BN']
                    sheet["C"+str(3+(8*(labelcount)))].font = Font(bold=True,size=10)
                    sheet["C"+str(3+(8*(labelcount)))].border = Border(bottom=thin)

                    ## EXPIRATION
                    if datetime.now().strftime('%A') == 'Friday':
                        day_modifier = 3
                    else:
                        day_modifier = 1 
                    
                    if type(val['expiration']) == str:
                        if 'months' in val['expiration']:
                            received_dt = datetime.now() + relativedelta(months = int(val['expiration'].split('months')[0]))
                            received_dt = received_dt + timedelta(days = day_modifier)
        
                        elif 'hours' in val['expiration']:
                            received_dt = datetime.now() + timedelta(hours = int(val['expiration'].split('hours')[0]))
                            received_dt = received_dt + timedelta(days = day_modifier)
                    else:                         
                        received_dt = datetime.now() + timedelta(days = int(val['expiration'])+day_modifier)
                    received_dt = received_dt.strftime('%d%b%y')
                        
                    sheet["A"+str(4+(8*(labelcount)))] = 'Expiration Date: ' + received_dt + '            PO#: ' + val['PO']
                    sheet["A"+str(4+(8*(labelcount)))].font = Font(bold=True,size=10)
                    sheet["A"+str(4+(8*(labelcount)))].border = Border(bottom=thin)

                    ## HAZARD INFO
                    sheet["C"+str(4+(8*(labelcount)))].alignment = Alignment(wrap_text = True, vertical = 'top',indent = 2.0)
                    if len(val['hazards']) > 340:
                        sheet["C"+str(4+(8*(labelcount)))].font = Font(size=7)
                    elif len(val['hazards']) > 210 and len(val['hazards']) <= 340:
                        sheet["C"+str(4+(8*(labelcount)))].font = Font(size=8)
                    else:
                        sheet["C"+str(4+(8*(labelcount)))].font = Font(size=9)
                    sheet["C"+str(4+(8*(labelcount)))] = val['hazards']
                    img = openpyxl.drawing.image.Image(r'hidden'+str(count)+'.png')
                    img.anchor = 'C'+str(7+(8*(labelcount)))
    
                    sheet.add_image(img)
    
                    ## STORAGE VESSELS ##
                    ###### IF BOTTLE PREP #######
                    sheet["A"+str(6+(8*(labelcount)))].alignment = Alignment(horizontal = 'left')
                    sheet["A"+str(6+(8*(labelcount)))].border = Border(bottom=thin)
                    if 'mL' in item[1]:
                        sheet["A"+str(6+(8*(labelcount)))] = 'Size: ' + item[1].split('mL')[0] + ' mL'
                        sheet["A"+str(6+(8*(labelcount)))].font = Font(bold=True,size=10)
                        dont_inc_vnum = 1
                    elif item[1] == '500':
                        sheet["A"+str(6+(8*(labelcount)))] = 'N:                     kg'
                        sheet["A"+str(6+(8*(labelcount)))].font = Font(bold=True,size=10)
                    elif len(item[1]) == 2 and item[1].endswith('L'):
                        sheet["A"+str(6+(8*(labelcount)))] = 'Size: ' + item[1].split('L')[0] + ' L'
                        sheet["A"+str(6+(8*(labelcount)))].font = Font(bold=True,size=10)
                        dont_inc_vnum = 1

                    elif len(item[1]) == 3 or len(item[1]) == 4 and item[1].endswith('L') and '500' not in item[1]:
                        sheet["A"+str(6+(8*(labelcount)))] = 'G:                 kg    T:                       kg    N:                     kg'
                        sheet["A"+str(6+(8*(labelcount)))].font = Font(bold=True,size=10)

                    elif item[1] == '500L':
                        sheet["A"+str(6+(8*(labelcount)))] = 'N:                     kg'
                        sheet["A"+str(6+(8*(labelcount)))].font = Font(bold=True,size=10)
                    
                    ## PH COND ETC ##    
                    phrow = ''
                    if val['pH'] == 1:
                        if 'stock' not in val['description'].lower():
                            phrow += 'pH:                 '
                    if val['cond'] == 1:
                        if 'stock' not in val['description'].lower():
                            phrow += 'Cond:                ms/Cm   '
                    phrow += val['temp'] +'°C'
                    if dont_inc_vnum != 1:
                        phrow += '       Vessel #'
                    if val['dark'] == 1:
                        phrow += '            DARK'
                    sheet["A"+str(5+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["A"+str(5+(8*(labelcount)))].alignment = Alignment(horizontal = 'left')
                    sheet["A"+str(5+(8*(labelcount)))] = phrow
                    sheet["A"+str(5+(8*(labelcount)))].font = Font(bold=True,size=9)

                    ## RECORDED BY
                    sheet["A"+str(7+(8*(labelcount)))].border = Border(bottom=thin)
                    sheet["A"+str(7+(8*(labelcount)))].font  = Font(b=True, color="000000",size=9)
                    sheet["A"+str(7+(8*(labelcount)))] = 'Recorded By/Date:                    Witnessed By/Date:'
                    sheet["A"+str(7+(8*(labelcount)))].alignment = Alignment(horizontal = 'left')
                    
                    img2 = openpyxl.drawing.image.Image(r'hidden')
                    img2.anchor = 'B'+str(7+(8*(labelcount)))
                    sheet.add_image(img2)

        if count_to_fpage >= 1:
            labelcount += count_to_fpage
        
        
    workbook.save(filename=r"hidden.xlsx")    
    
    totallabels = totalnumofvessels + totaltypesofvessels
    

def ignition(prid,password):
    i = 0
    while i == 0:
        if pag.locateCenterOnScreen(ignition_open): # If ignition is open, maximize it
            print('already open')
            xloc,yloc = pag.locateCenterOnScreen(ignition_open)
            pag.click(xloc,yloc)
        else: 
            print('closed')
            if pag.locateCenterOnScreen(ignition_closed, confidence=0.9): # If ignition is closed, open it
                x,y = pag.locateCenterOnScreen(ignition_closed, confidence=0.9)
                pag.click(x,y,button='right')
            while not pag.locateCenterOnScreen(ignition_rclick, confidence=0.9):
                pass
            x,y = pag.locateCenterOnScreen(ignition_rclick, confidence=0.9)
            pag.click(x,y)
            while not pag.locateCenterOnScreen(ignition_logon, confidence=0.9): 
                print('waiting')

        if pag.locateCenterOnScreen(ignition_logon, confidence=0.9):

            pag.typewrite(prid); pag.hotkey('tab')
            pag.typewrite(password); pag.hotkey('tab')
            pag.hotkey('enter'); time.sleep(3)
        while not pag.locateCenterOnScreen(ignition_reporting, confidence=0.9):
            pass
        if pag.locateCenterOnScreen(ignition_reporting, confidence=0.9):
            pag.click(953,571); time.sleep(0.5); 
            i = 1
        else:
            pag.click(x=1004, y=540)
            i = 1

    if i == 1:
        pag.click(433,478) # (x=486, y=484)
        
def make_reports(pns_to_lookup_prep):
    """
    Make ignition reports for each prep if they require them.
    """
    time.sleep(2)

    searchbyrecipe = (595, 466)
    addidentifier = (713, 401)
    identifier = (353, 400)
    equipment = (508, 512)
    recipe = (772, 572)
    batchID = (780, 620)
    addbatchID = (537,686)
    rec_params = (497, 816)
    rep_values = (500, 855)
    recipe_1 = (1152, 363)
    recipe_2 = (1204, 394)
    recipe_3 = (1256, 425)
    complete_report = (1473, 973)
    continue_to_sip = 0
    submit_report = 0
            
    def create_identifier(val):
        """
        Create an identifier in ignition for a prep.
        """
        to_write = val['SAP PN']+"_"+val['BN']

        pag.click(392,348)
        pag.typewrite('s'); pag.hotkey('enter'); time.sleep(0.5);
        pag.click(714,399); pag.click(383,396); pag.typewrite(to_write); time.sleep(1); pag.hotkey('enter'); time.sleep(0.5)

        if pag.locateCenterOnScreen(ignition_alreadymade):
            pag.hotkey('enter')
            pag.click(392,348); pag.hotkey('tab')
            pag.typewrite(to_write)
        else:
            pag.hotkey('enter')
        time.sleep(2)
        
    def verify_cip(rec_1_copy):
        """
        Verify our CIP is within expiration date.
        """
        dt_object = rec_1_copy.split('-')[1]+" "+rec_1_copy.split('-')[2] 
        dt_obj2 = datetime.strptime(dt_object,'%d%b%Y %H:%M:%S')
        received_dt = dt_obj2 + timedelta(days = 13)
        time_on_cip_exp = received_dt - datetime.now()

        return time_on_cip_exp.total_seconds()
    
    def preptank_portables(pns_to_lookup_prep): 
        """
        Add portable tank CIPS to our report.
        """
        ptank_holder = []
        count = 1
        time.sleep(3)
        tracker = {}
        tracker2 = {}
        batch_holder = []
        recipe_1 = (1152, 363)
        recipe_2 = (1204, 394)
        recipe_3 = (1256, 425)
        checks = ''

        for key,val in pns_to_lookup_prep.items(): # Create temp list of all portables needed
            for item in val['Storage']:
                if item in ['1128','1132','1186','2150','1520']:
                    ptank_holder.append(item)
        for item in ptank_holder: 
            tracker[item] = False
            
        if not pag.locateCenterOnScreen(ignition_searchbyrecipe_checked, confidence=0.8):
            pag.click(searchbyrecipe)
            
        while not all(value == True for value in tracker.values()): # Iterate through until we have the required/passing CIPS

            pag.click(351, 511)
            pag.typewrite('UP2171'); pag.hotkey('enter'); 

            while not pag.locateCenterOnScreen(ignition_recipeready, confidence=0.8): # Wait for recipe to load
                pass
            x,y = pag.locateCenterOnScreen(ignition_recipeready, confidence=0.8)
            pag.click(x,y)
            pag.typewrite('UP2171_CIP_217')
            pag.hotkey('tab');
            while not pag.locateCenterOnScreen(ignition_batchidready, confidence=0.8):
                pass
            x,y = pag.locateCenterOnScreen(ignition_batchidready, confidence=0.8)
            pag.click(x,y)
            pag.hotkey('down'); pag.hotkey('tab'); pag.hotkey('space'); 
            for i in range(9):
                pag.click(459, 613)
                pag.hotkey('down'); pag.hotkey('tab'); pag.hotkey('space')

            # Add batch ID
            x,y = pag.locateCenterOnScreen(ignition_previewreport, confidence=0.8)
            pag.click(x,y)
            time.sleep(30)
            pag.click(recipe_2,button='right')
            time.sleep(1)
            pag.click(1210,420)
            ### CHANGE TO SAVE TO LOCAL DRIVE FOR SPEED
            count_str = str(count)
            total_dir = r"hidden"+count_str+".pdf"
            pag.typewrite(total_dir); pag.hotkey('enter')
            time.sleep(25)

            reader = PdfReader(total_dir) # Make pdf and check for our tank
            number_of_pages = len(reader.pages)
            all_pages = []
            for i in reader.pages:
                text = i.extract_text()
                all_pages.append(text)

            all_pages.pop(0)
            for elem in ptank_holder:
                for item in all_pages:
                    new_elem = 'TK'+str(elem)
                    if new_elem in item and 'CONDUCTIVITY_TEST_PASSED PASS' in item:
                        batch_id = item.split('BatchID - ')[1].split('\n')[0]
                        tracker[elem] = True
                        tracker2[elem] = {'Status':'PASS', 'Batch ID':batch_id}

        for key,val in tracker2.items():
            batch_holder.append(val['Batch ID'])
        
        i = 0
        while i <= len(batch_holder):
            is_in = 0
            pag.click(1565,362+24*i);time.sleep(0.2) # Click BatchID 1
            pag.hotkey('ctrl','c')
            checks = pyperclip.paste()
            for item in batch_holder:
                if item in str(checks):
                    is_in = 1

            if is_in == 0 and 'UP2171' in checks:
                pag.click(1565,362+24*i); pag.click(1565,362+24*i,button='right'); time.sleep(0.2) 
                pag.click(1570,368+24*i); pag.hotkey('enter'); time.sleep(0.1)
            else:
                i += 1
    
    def add_CIPS_to_report(val,num_portables,hold_vessel):
        """
        Add CIPS to report.
        """
        print('Adding CIPS')
        recipe_1_loc = (1372,366)
        cip_count = -1
        if val['CIP'] == ['any']:
            if pag.locateCenterOnScreen(ignition_searchbyrecipe_checked, confidence=0.8): # Turn off search by recipe
                pag.click(searchbyrecipe)
        else:
            if not pag.locateCenterOnScreen(ignition_searchbyrecipe_checked, confidence=0.8):
                pag.click(searchbyrecipe)

        for cipss in val['CIP']:
            cip_count += 1
            recipe_index = cip_count+num_portables
            pag.click(equipment); time.sleep(0.5) # Click on 'Equipment'
            if cip_count >= 1:
                for item in val['Storage']:
                    if cips_hold[item][val['Area']] == val['CIP'][cip_count]:
                        hold_vessel = item
                pag.typewrite('TK'+hold_vessel); pag.hotkey('enter')
            else:       
                pag.typewrite('TK'+val['Prep Vessel']); pag.hotkey('enter');

            if val['CIP'] == ['any']: # TANK TO BAGS DON'T NEED RECIPE
                while not pag.locateCenterOnScreen(ignition_batchidready, confidence=0.8):
                    pass
                if pag.locateCenterOnScreen(ignition_batchidready, confidence=0.8):
                    x,y = pag.locateCenterOnScreen(ignition_batchidready, confidence=0.8)
                    pag.click(x,y)
                elif pag.locateCenterOnScreen(ignition_batchidready2, confidence=0.8):
                    x,y = pag.locateCenterOnScreen(ignition_batchidready2, confidence=0.8)
                    pag.click(x,y)
            else: # NOT TANK TO BAGS WE NEED RECIPE
                while not pag.locateCenterOnScreen(ignition_recipeready, confidence=0.8):
                    pass
                x,y = pag.locateCenterOnScreen(ignition_recipeready, confidence=0.8)
                pag.click(x,y)
                pag.typewrite('TK'+val['Prep Vessel']+'_CIP_'+val['CIP'][cip_count])
                pag.hotkey('tab');
                while not pag.locateCenterOnScreen(ignition_batchidready, confidence=0.8):
                    pass
                if pag.locateCenterOnScreen(ignition_batchidready, confidence=0.8):
                    x,y = pag.locateCenterOnScreen(ignition_batchidready, confidence=0.8)
                    pag.click(x,y)
                elif pag.locateCenterOnScreen(ignition_batchidready2, confidence=0.8):
                    x,y = pag.locateCenterOnScreen(ignition_batchidready2, confidence=0.8)
                    pag.click(x,y)

            pag.hotkey('down'); pag.hotkey('tab'); pag.hotkey('space'); time.sleep(0.5)
            
            pag.click(1372,366+(28*recipe_index)); pag.hotkey('ctrl','c')
            rec_cip_copy = pyperclip.paste()
            if verify_cip(rec_cip_copy) > 50000: # Verify our CIP is within expiration
                continue
            else:
                if val['CIP'] == ['any']:
                    print('Latest CIP is expired, breaking...')
                    break
                else:
                    print('Latest CIP is expired, try the rinse...')
                    pag.click(1565,362+24*recipe_index); pag.click(1565,362+24*recipe_index,button='right'); time.sleep(0.2) 
                    pag.click(1570,368+24*recipe_index); pag.hotkey('enter'); time.sleep(0.1)
                    pag.click(recipe)
                    pag.typewrite('TK'+val['Prep Vessel']+'_CIP_'+val['CIP'][cip_count]+'_RINSE'); pag.hotkey('tab')
                    while not pag.locateCenterOnScreen(ignition_batchidready, confidence=0.8):
                        pass
                    x,y = pag.locateCenterOnScreen(ignition_batchidready)
                    pag.click(x,y); pag.hotkey('down'); pag.hotkey('tab'); pag.hotkey('space'); time.sleep(0.5)
                    pag.click(1372,366+(28*recipe_index)); pag.hotkey('ctrl','c')
                    rec_rinse_copy = pyperclip.paste()
                    if verify_cip(rec_rinse_copy) > 50000:
                        continue
                    else:
                        print('Latest CIP and Rinse expired, breaking...')
                        break
              
    def add_SIPS_to_report(val,num_loc): 
        """
        Add SIPS to our report and verify they are not expired
        """
        recipe_1_loc = (1372,366)
        for item in val['Storage']:
            if sips[item][val['Area']] == val['SIP'][0]:
                hold_vessel = item
        if not pag.locateCenterOnScreen(ignition_searchbyrecipe_checked, confidence=0.8):
            pag.click(searchbyrecipe)
        if hold_vessel == '1523' or hold_vessel == '1526' or hold_vessel == '2149' or hold_vessel == '2150':
            pag.click(recipe)
            pag.hotkey('down'); pag.hotkey('down'); pag.hotkey('down'); pag.hotkey('tab'); time.sleep(10)
        elif hold_vessel == '1128' or hold_vessel == '1132' or hold_vessel == '1186' or hold_vessel == '1028' or hold_vessel == '1205' or hold_vessel == '1926' or hold_vessel == '1927':
            pag.click(recipe)
            pag.hotkey('down'); pag.hotkey('down'); pag.hotkey('down'); pag.hotkey('down'); pag.hotkey('tab'); time.sleep(10)
        else:
            pag.click(recipe)
            pag.typewrite('TK'+hold_vessel+'_SIP_'+val['SIP'][0])
            pag.hotkey('tab'); time.sleep(10)
            pag.click(batchID); time.sleep(0.5)
            pag.hotkey('down'); pag.hotkey('tab'); pag.hotkey('space');time.sleep(0.1) # Add batch ID
        pag.click(1372,366+(28*num_loc))
        pag.hotkey('ctrl','c')
        rec_sip_copy = pyperclip.paste()
        if verify_cip(rec_sip_copy) < 50000:
            print('Latest SIP expired, breaking...')
            return False
        else:
            return True
            
    def create_print_return(val,num_portables):
        """
        Finalize our report, print it, and return to create another report.
        """
        pag.click(complete_report); pag.hotkey('enter')
        time.sleep(2); pag.click(185,86); time.sleep(1); pag.click(720, 837)
        time.sleep(3); pag.click(872, 394); pag.typewrite('s'); pag.hotkey('tab'); time.sleep(1)
        pag.typewrite(val['SAP PN']+"_"+val['BN']); time.sleep(1);pag.hotkey('enter'); time.sleep(1)
        pag.click(886, 599) # Full report
        if num_portables >= 1:
            pass
        else:
            pag.hotkey('tab'); pag.hotkey('space'); pag.hotkey('tab'); pag.hotkey('space')
        xp,yp = pag.locateCenterOnScreen(ignition_viewreport, confidence = 0.8)
        pag.click(xp,yp)
        while not pag.locateCenterOnScreen(ignition_reportsummary, confidence = 0.8):
            pass

        pag.click(951, 827,button='right')
        while not pag.locateCenterOnScreen(ignition_saveaspdf, confidence = 0.8):
            pass

        x,y = pag.locateCenterOnScreen(ignition_saveaspdf, confidence = 0.8)
        pag.click(x,y)
        pag.typewrite(val['SAP PN'] + '_' + val['BN'] + '_report.pdf')
        pag.hotkey('enter')
        time.sleep(1)
        xb,yb = pag.locateCenterOnScreen(ignition_back, confidence = 0.8)
        pag.click(xb,yb)
        while not pag.locateCenterOnScreen(ignition_home, confidence = 0.8):
            pass

        xh,yh = pag.locateCenterOnScreen(ignition_home, confidence = 0.8)
        pag.click(xh,yh)
                
    for key,val in pns_to_lookup_prep.items():
        num_portables = 0
        hold_vessel = ''
        num_loc = num_portables + len(val['CIP'])
        create_identifier(val) # Create our identifier for each prep
        
        for item in val['Storage']: # Add portable tank recipes if we have any
            if item in ['1128','1132','1186','2150','1520']:
                num_portables += 1
        if num_portables >= 1:
            preptank_portables(pns_to_lookup_prep) ########### FIX #@###@$#@$#$#
        
        add_CIPS_to_report(val,num_portables,hold_vessel) # Add CIPS to report
        if len(val['SIP']) >= 1:
            sip_check = add_SIPS_to_report(val,num_loc)
            if sip_check == False:
                break
            else:
                create_print_return(val,num_portables)
        else:
            create_print_return(val,num_portables)

layout = [[sg.Text('Prid:         ',font=('Arial',11)), sg.Input('',key='-pridinput-')],
          [sg.Text('Password:',font=('Arial',11)), sg.Input('',key='-passinput-')],
          [sg.Button('Run', font=('Arial',11)),sg.Exit()]]
layout2 = [[sg.Text('',font=('Arial',11),key='-update-'),sg.Exit()]]
layout4 = [[sg.Text('',font=('Arial',11),key='-update2-'),sg.Exit()]]

window = sg.Window('AutoLabels',layout,
                   keep_on_top=True,
                   transparent_color='red',
                   alpha_channel = 1.0,
                   background_color = 'red',
                   no_titlebar = True, finalize=True,location=(1400,0))
   
while True:             # Event Loop`
    event, values = window.read()
    correct_preps = {}
    
    if event in (None, 'Exit'):
        break
    
    elif event == 'Run':
        window.close()
        prid = values['-pridinput-']
        password = values['-passinput-']
        window = sg.Window('AutoLabels',layout2,
                   keep_on_top=True,
                   transparent_color='red',
                   alpha_channel     = 1.0,
                   background_color = 'red',
                   no_titlebar = True, finalize=True,location=(1400,0))
        window['-update-'].update('Getting bearer token from EQV')
        window.refresh()
        bearer = get_bearer()
        window['-update-'].update('Sending email to Power Automate')
        window.refresh()
        preplist = sendandwait()
        window['-update-'].update('Planner extracted')
        window.refresh()
        time.sleep(0.5)
        window['-update-'].update('Extracting PN information from planner')
        window.refresh()
        pns_to_lookup_prep0 = extractplanner(preplist,bearer)
        
        window['-update-'].update('Getting SPR veeva ID to extract SPR information')
        window.refresh()
        sprveevaid = getsprid(pns_to_lookup_prep0,bearer) 
        window['-update-'].update('Decoding SPR')
        window.refresh()
        decodedsprs = decodespr(sprveevaid,bearer,pns_to_lookup_prep0)
        window['-update-'].update('Compiling information')
        window.refresh()
        window['-update-'].update('Opening SAP')
        window.refresh()
        final_openSAP()
        window['-update-'].update('Going to transaction COOISPI')
        window.refresh()
        goto_COOISPI()

        window['-update-'].update('Gathering BN/PO from SAP')
        window.refresh()
        pns_to_lookup_prepc = COOISPI_2(pns_to_lookup_prep0)
        for key,val in pns_to_lookup_prepc.items():
            for key2,val2 in sprveevaid.items():
                if val['PN'] in key2:
                    if val2[0] == ' ': 
                        pns_to_lookup_prepc[key]['description'] = val2[1:].split('@')[0]
                    else:
                        pns_to_lookup_prepc[key]['description'] = val2.split('@')[0]

            for key3,val3 in decodedsprs.items():
                if val['PN'] == val3['PN']:
                    pns_to_lookup_prepc[key]['pH'] = val3['pH']
                    pns_to_lookup_prepc[key]['turb'] = val3['turb']
                    pns_to_lookup_prepc[key]['cond'] = val3['cond']
                    pns_to_lookup_prepc[key]['osmo'] = val3['osmo']
                    pns_to_lookup_prepc[key]['dark'] = val3['dark']
                    pns_to_lookup_prepc[key]['expiration'] = val3['expiration']
                    pns_to_lookup_prepc[key]['hazards'] = val3['hazards']
                    pns_to_lookup_prepc[key]['temp'] = val3['temp']
        print('pns',pns_to_lookup_prepc)
        window['-update-'].update('Creating labels')
        window.refresh()
        createlabelstart(pns_to_lookup_prepc)
        window.close()
        layout3 = [[sg.Text('Correct any wrong information if any to remake labels')],
          [[sg.Text(pns_to_lookup_prepc[i]['SAP PN'],key='pn'+str(i)),sg.Input(pns_to_lookup_prepc[i]['BN'],key='bninput'+str(i)),sg.Input(pns_to_lookup_prepc[i]['PO'],key='poinput'+str(i))] for i in range(len(pns_to_lookup_prepc))],
          [sg.Exit(),sg.Button('Remake labels'),sg.Button('Continue')]]
        window = sg.Window('AutoLabels',layout3,
                   keep_on_top=True,
                   transparent_color='red',
                   alpha_channel     = 1.0,
                   background_color = 'white',
                   no_titlebar = True, finalize=True,location=(1200,0))
    elif event == 'Remake labels':
        for keyc,valc in pns_to_lookup_prepc.items():
            new_BN = values['bninput'+str(keyc)]
            new_PO = values['poinput'+str(keyc)]
            correct_preps[keyc] = valc
            correct_preps[keyc]['BN'] = new_BN
            correct_preps[keyc]['PO'] = new_PO
        createlabelstart(correct_preps)
        pns_to_lookup_prepc = correct_preps
    elif event == 'Continue':
        window.close()
        window = sg.Window('AutoLabels',layout4,
                   keep_on_top=True,
                   transparent_color='red',
                   alpha_channel     = 1.0,
                   background_color = 'red',
                   no_titlebar = True, finalize=True,location=(1400,0))
        window['-update2-'].update('Checking if reports need to be made')
        window.refresh()
        reports_to_make = {}
        for key,val in pns_to_lookup_prepc.items():
            if val['CIP'] != '':
                reports_to_make[key] = val
        if len(reports_to_make) >= 1:
            window['-update2-'].update('Logging in to ignition')
            window.refresh()
            ignition(prid,password)

            window['-update2-'].update('Generating reports')
            window.refresh()
            make_reports(reports_to_make) 
            window['-update2-'].update('Process complete')
window.close()